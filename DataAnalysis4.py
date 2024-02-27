import tkinter.filedialog
from tkinter import messagebox
from tkinter import *  # 导入tkinter模块【必要步骤】
import xlrd
import xlwt
import re
import openpyxl
from openpyxl.styles import PatternFill


def FileOpen():
#文件对话框,能打开写死的文件框
    return_value = tkinter.filedialog.askopenfilename()
    if return_value.strip()!=" ":
        filename1.set(return_value)
    else:
        print("未上传所需文件")
    #直接打开文件选择框
    print(type(return_value),return_value)
def FileOpen2():
#文件对话框,能打开写死的文件框
    return_value = tkinter.filedialog.askopenfilename()
    if return_value.strip()!=" ":
        filename2.set(return_value)
    else:
        print("未上传所需文件")
    #直接打开文件选择框
    print(type(return_value),return_value)

def FileOpen21():
#文件对话框,能打开写死的文件框
    return_value = tkinter.filedialog.askopenfilename()
    if return_value.strip()!=" ":
        filename21.set(return_value)
    else:
        print("未上传所需文件")
    #直接打开文件选择框
    print(type(return_value),return_value)
def FileSave():
    #设置保存文件，并返回文件名，指定文件名后缀为“.xls
    r=tkinter.filedialog.asksaveasfilename(title="保存文件",defaultextension='.xls'
                                           # initialdir="E:\Tools\files",
                                           # initialfile='test.py'
                                           )
    if r.strip()!=' ':
        filewname.set(r)
    else:
        print("未选择需要保存的文件")
    print(r)

def FileSave2():
    #设置保存文件，并返回文件名，指定文件名后缀为“.xls
    r=tkinter.filedialog.asksaveasfilename(title="保存文件",defaultextension='.xls'
                                           # initialdir="E:\Tools\files",
                                           # initialfile='test.py'
                                           )
    if r.strip()!=' ':
        filewname2.set(r)
    else:
        print("未选择需要保存的文件")
    print(r)


root = Tk()  # 创建主窗口【必要步骤】
# 将该窗口赋值给root变量，方便后续使用
root.title('纪委数据模型分析软件')
root.geometry('960x480+150+100')#设置窗口大小及位置
root.wm_attributes('-alpha')#设置透明度为0.7
root.resizable(1,1)#窗口大小不可改变两个布尔值分别代表窗口的长和宽是否可改变（可以用0和1代替布尔值）

filename1 = tkinter.StringVar()
filename2=tkinter.StringVar()
filewname=tkinter.StringVar()

filename21 = tkinter.StringVar()
filewname2=tkinter.StringVar()

# toplevel=Toplevel(root)
# toplevel.title('子窗口')
#打开文件
#Entry：文本框输入框，Label：标签熟悉
def ui():
    canvas = tkinter.Canvas(root, bg='green', highlightthickness=0)
    canvas.grid(row=0, column=0, columnspan=4, rowspan=7, sticky='nsew', padx=5, pady=5)

    # 在Canvas上绘制一个矩形（这里的坐标和大小需要根据实际情况调整）
    rect = canvas.create_rectangle(0, 0, 440, 260, outline='black', width=2)

    canvas2 = tkinter.Canvas(root, bg='yellow', highlightthickness=0)
    canvas2.grid(row=0, column=5, columnspan=4, rowspan=7, sticky='nsew', padx=5, pady=5)

    # 在Canvas上绘制一个矩形（这里的坐标和大小需要根据实际情况调整）
    rect2 = canvas2.create_rectangle(0, 0, 440, 260, outline='black', width=2)


    tkinter.Label(root, text='供应商库').grid(row=1, column=1, padx=5, pady=7)

    #第一个输入框
    filename1path=tkinter.Entry(root, textvariable=filename1, width=40 )
    filename1path.grid(row=1, column=2, padx=5, pady=5)
    b1=Button(root,relief='groove',activebackground='pink',bg='lightblue',overrelief='ridge',text='上传文件',command=FileOpen).grid(row=1, column=3, padx=5, pady=5)#创建按钮
    # b1.pack(side="left")

    tkinter.Label(root, text='有异议供应商').grid(row=2, column=1, padx=5, pady=5)
    #第二个输入框
    filename2path=tkinter.Entry(root, textvariable=filename2, width=40 )
    filename2path.grid(row=2, column=2, padx=5, pady=5)
    b2=Button(root,relief='groove',activebackground='pink',bg='lightblue',overrelief='ridge',text='上传文件',command=FileOpen2).grid(row=2, column=3, padx=5, pady=5)
    # b2.pack(side="left")

    # b.place(width=100,height=30,x=150,y=200)#防止按钮
    tkinter.Label(root, text='保存文件').grid(row=5, column=1, padx=5, pady=5)
    #第三个输入框
    filewnamepath=tkinter.Entry(root, textvariable=filewname, width=40 )
    filewnamepath.grid(row=5, column=2, padx=5, pady=5)
    b3=Button(root,text='保存文件',command=FileSave,relief='groove',activebackground='pink',bg='orange',overrelief='ridge',).grid(row=5, column=3, padx=5, pady=5)
    # b3.pack(side="left")

    tkinter.Button(root,text="开始对比",activebackground='pink',bg='lightblue',overrelief='ridge',command=lambda :contrast_button_clicked(filename1path.get(),filename2path.get(),filewnamepath.get())).grid(row=6, column=3, padx=5, pady=5)



    tkinter.Label(root, text='数据提取').grid(row=2, column=5, padx=5, pady=7)
    # 第二部分第一个输入框
    filename21path = tkinter.Entry(root, textvariable=filename21, width=40)
    filename21path.grid(row=2, column=7, padx=5, pady=5)
    b21 = Button(root, relief='groove', activebackground='pink', bg='lightblue', overrelief='ridge', text='上传文件',
                command=FileOpen21).grid(row=2, column=8, padx=5, pady=5)  # 创建按钮
    # b1.pack(side="left")


    tkinter.Label(root, text='保存文件').grid(row=3, column=5, padx=5, pady=5)
    # 第二部分第二个输入框
    filewnamepath2 = tkinter.Entry(root, textvariable=filewname2, width=40)
    filewnamepath2.grid(row=3, column=7, padx=5, pady=5)
    b22 = Button(root, text='保存文件', command=FileSave2, relief='groove', activebackground='pink', bg='orange',
                overrelief='ridge', ).grid(row=3, column=8, padx=5, pady=5)
    # b22.pack(side="left")


    tkinter.Button(root, text="开始提取", activebackground='pink', bg='lightblue', overrelief='ridge',
                   command=lambda: contrast_buttontiqu_clicked(filename21path.get(),
                                                           filewnamepath2.get())).grid(row=6, column=8, padx=5, pady=5)



    #窗口主循环
    root.mainloop()  # 主窗口进入消息事件循环【必要步骤】放在最后



def contrast_buttontiqu_clicked(filename21path,  filewnamepath):
    flag = dataprocess(filename21path,filewnamepath)
    if flag:
        tkinter.messagebox.showinfo("提示", "对比成功！共有"+str(match)+"组数据重合，"+"共有" +str(diff)+"组数据不一致，详细情况请到"+ filewnamepath + "文件查看")
    else:
        tkinter.messagebox.showinfo("提示", "对比失败！请检查输入路径是否正确")

def contrast_button_clicked(filename1path, filename2path, filewnamepath):
    match, diff,flag = main(filename1path, filename2path, filewnamepath)
    if flag:
        tkinter.messagebox.showinfo("提示", "对比成功！共有"+str(match)+"组数据重合，"+"共有" +str(diff)+"组数据不一致，详细情况请到"+ filewnamepath + "文件查看")
    else:
        tkinter.messagebox.showinfo("提示", "对比失败！请检查输入路径是否正确")




def readexcel(file):
    # 打开Excel文件
    workbook = openpyxl.load_workbook(file)
    # 选择指定的Sheet
    sheet = workbook['Sheet1']
    # 选择需要获取的列的列号
    col_numbers = [2]
    # 列名
    titles = ['供应商']
    result = []
    # 从第2行开始获取数据
    for i in range(2, sheet.max_row + 1):
        # 每一行的数据
        row_data = []
        for col in col_numbers:
            row_data.append(sheet.cell(row=i, column=col).value)
        # 构建字典
        row_dict = dict(zip(titles, row_data))
        result.append(row_dict)
    return result



def comparefile(list1, list2):
    list1 = tuple(tuple(item.items()) for item in list1)
    list2 = tuple(tuple(item.items()) for item in list2)
    match = len(set(list1) & set(list2))  # 记录重合数
    diff = len(set(list1) - set(list2)) + len(set(list2) - set(list1))  # 记录差异数
    match1 = (set(list1) & set(list2))  # 记录重合元素
    differ1 = set(list1) - set(list2)  # 记录list1的差异元素
    differ2 = set(list2) - set(list1)  # 记录list2的差异元素
    differ1_dicts = [dict(item) for item in list(differ1)]
    differ2_dicts = [dict(item) for item in list(differ2)]
    match1_dicts = [dict(item) for item in list(match1)]

    print("1111", differ1_dicts, differ2_dicts)
    print(f"重合数：{match}，差异数：{diff}")
    # differ1=list(differ1)
    # differ2=list(differ2)
    return differ1_dicts, differ2_dicts, match1_dicts, match, diff



def is_valid_company_name(name):
    return not re.match(r'标段\s*\d+：', name) and not re.match(r'\d+、', name)



def dataprocessfile(list21):
    list21 = tuple(tuple(item.items()) for item in list21)
    matched_texts = []
    pattern_loose = r'[\u4e00-\u9fa5]+(?:\（[\u4e00-\u9fa5]+?\）)?[\u4e00-\u9fa5]*(?:有限公司|股份公司|物流公司|运输公司|集团|工程公司|环保工程|机械设备|海洋工程|科技发展|油气设备等)?'

    # 使用正则表达式的findall方法来查找所有可能的匹配项

    for list in list21:
        # 使用正则表达式的findall方法来查找所有可能的匹配项
        matches = re.findall(pattern_loose, list)

        # 如果找到了匹配项，将原始文本添加到matched_texts列表中
        if matches:
            matched_texts.append({"original_text": list, "companies": matches})




    #filtered_matches = [match for match in matches_loose if is_valid_company_name(match)]
    for match in matched_texts:
        print("公司名称:", match.strip())

    differ21 = set(list21)
    differ21_dicts = [dict(item) for item in list(differ21)]

    return differ21_dicts

def writeexcel2(differ21,filewname):
    global flag
    # 创建一个新的Excel工作簿
    wb = openpyxl.Workbook()
    # 获取正在使用的工作表
    sheet21 = wb.active
    sheet21.title = '数据处理'


    # 写入match1数据到sheet1
    if differ21:
        keys = differ21[0].keys()
        # 写入表头
        for col_num, key in enumerate(keys, start=1):
            col_letter = openpyxl.utils.get_column_letter(col_num)
            sheet21[f"{col_letter}1"] = key
            # 写入数据行
        for row_num, row_data in enumerate(differ21, start=2):
            for col_num, key in enumerate(keys, start=1):
                col_letter = openpyxl.utils.get_column_letter(col_num)
                sheet21[f"{col_letter}{row_num}"] = row_data.get(key, "")
                # 为当前行的所有单元格添加淡蓝色填充（如果需要）
            # 注意：这里我注释掉了，因为通常不会这么做，除非有特殊需求
            # for col_num in range(1, len(keys) + 1):
            #     col_letter = openpyxl.utils.get_column_letter(col_num)
            #     sheet1[f"{col_letter}{row_num}"].fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
    else:
        print("没有数据")        # 写入differ2数据到sheet2
            # 保存Excel工作簿
    wb.save(filewname)
    flag = True
    print("保存数据到excel文件完成")
    return flag
def writeexcel1(match1, differ1,differ2,filewname):
    global flag
    # 创建一个新的Excel工作簿
    wb = openpyxl.Workbook()
    # 获取正在使用的工作表
    sheet1 = wb.active
    sheet1.title = '两表相同数据'
    sheet2 = wb.create_sheet(title='表1差值数据')
    sheet3= wb.create_sheet(title='表2差值数据')

    # 写入match1数据到sheet1
    if match1:
        keys = match1[0].keys()
        # 写入表头
        for col_num, key in enumerate(keys, start=1):
            col_letter = openpyxl.utils.get_column_letter(col_num)
            sheet1[f"{col_letter}1"] = key
            # 写入数据行
        for row_num, row_data in enumerate(match1, start=2):
            for col_num, key in enumerate(keys, start=1):
                col_letter = openpyxl.utils.get_column_letter(col_num)
                sheet1[f"{col_letter}{row_num}"] = row_data.get(key, "")
                # 为当前行的所有单元格添加淡蓝色填充（如果需要）
            # 注意：这里我注释掉了，因为通常不会这么做，除非有特殊需求
            # for col_num in range(1, len(keys) + 1):
            #     col_letter = openpyxl.utils.get_column_letter(col_num)
            #     sheet1[f"{col_letter}{row_num}"].fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
    else:
        print("match1没有数据")
        # 写入differ2数据到sheet2
    if differ1:
        keys = differ1[0].keys()
        # 写入表头
        for col_num, key in enumerate(keys, start=1):
            col_letter = openpyxl.utils.get_column_letter(col_num)
            sheet2[f"{col_letter}1"] = key
            # 写入数据行
        for row_num, row_data in enumerate(differ1, start=2):
            for col_num, key in enumerate(keys, start=1):
                col_letter = openpyxl.utils.get_column_letter(col_num)
                sheet2[f"{col_letter}{row_num}"] = row_data.get(key, "")
                # 同上，注释掉了填充颜色的部分
    else:
        print("differ1没有数据")

        # 写入differ2数据到sheet2
    if differ2:
        keys = differ2[0].keys()
        # 写入表头
        for col_num, key in enumerate(keys, start=1):
            col_letter = openpyxl.utils.get_column_letter(col_num)
            sheet3[f"{col_letter}1"] = key
            # 写入数据行
        for row_num, row_data in enumerate(differ2, start=2):
            for col_num, key in enumerate(keys, start=1):
                col_letter = openpyxl.utils.get_column_letter(col_num)
                sheet3[f"{col_letter}{row_num}"] = row_data.get(key, "")
                # 同上，注释掉了填充颜色的部分
    else:
        print("differ2没有数据")

        # 保存Excel工作簿
    wb.save(filewname)
    flag = True
    print("保存差异值到excel文件完成")
    return flag

def dataprocess(filename21path,filewnamepath):
    # file1 = filename1.get()
    # file2 = filename2.get()
    # resultfile = filewname.get()
    list21 = readexcel(filename21path)

    print(list21)
    differ21 = dataprocessfile(list21)
    flag=writeexcel2(differ21,filewnamepath)
    # print("flag",flag)
    return flag

def main(filename1path,filename2path,filewnamepath):
    # file1 = filename1.get()
    # file2 = filename2.get()
    # resultfile = filewname.get()
    list1 = readexcel(filename1path)
    list2 = readexcel(filename2path)
    print(list1)
    print(list2)
    differ1, differ2,match1,match,diff = comparefile(list1, list2)
    flag=writeexcel1(match1,differ1, differ2,filewnamepath)
    # print("flag",flag)
    return match,diff,flag


if __name__ == '__main__':
    ui()